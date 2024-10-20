from openpyxl import load_workbook

def is_float(v):
    try:
        f=float(v)
    except ValueError:
        return False
    return True

wb = load_workbook(filename = 'Apollo_Rx_Latency_Calculator_B0.xlsm')

for s in wb.sheetnames:
    print("tab = ", "\"", s, "\"", sep="")

    ws = wb[s]
    for row in ws:
        # ['A2':'B5']:
        for c in row:
            if c.value is not None:
                sheet = s
                if " " in s:
                    sheet = "'" + sheet + "'"
                if is_float(c.value):
                    print(sheet, "!", c.coordinate, "=", c.value, sep="")
                elif c.value.startswith("="):
                    val = c.value.replace("_xlfn.FLOOR.MATH","FLOOR_MATH")
                    print(sheet, "!", c.coordinate, "=", val[1:], sep="")
                else:
                    print(sheet, "!", c.coordinate, " = \"", c.value, "\"", sep="")

# print("")
# print("defined names :")
for k, v in wb.defined_names.items():
    print(k, "alias", v.attr_text.replace('$',''))
