from openpyxl import load_workbook

def is_float(v):
    try:
        f=float(v)
    except ValueError:
        return False
    return True

wb = load_workbook(filename = 'Apollo_Rx_Latency_Calculator_B0.xlsm')

print (wb.sheetnames)
for s in wb.sheetnames:
    print("\nSheet", s, ":")

    ws = wb[s]
    for row in ws:
        # ['A2':'B5']:
        for c in row:
            if c.value is not None:
                if is_float(c.value) or c.value.startswith("="):
                    print(c.coordinate, "=", c.value)
                else:
                    print(c.coordinate, " = \"", c.value, "\"", sep="")

print("")
print("defined names :")
for k, v in wb.defined_names.items():
    print(k, "->", v.attr_text.replace('$',''))
